from openpyxl import load_workbook

from app.rag.parsers.base_parser import BaseParser


class ExcelParser(BaseParser):

    def parse(
        self,
        file_path: str,
    ) -> str:

        workbook = load_workbook(
            file_path,
            data_only=True,
        )

        rows = []

        for sheet in workbook.worksheets:

            for row in sheet.iter_rows(values_only=True):

                values = [
                    str(cell)
                    for cell in row
                    if cell is not None
                ]

                if values:

                    rows.append(
                        " | ".join(values)
                    )

        return "\n".join(rows)